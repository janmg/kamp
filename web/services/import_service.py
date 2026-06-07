"""CSV and Excel import helpers shared by CLI and web flows."""

from __future__ import annotations

import csv
import io
import random
import re
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import TextIO

from sqlalchemy.exc import IntegrityError

from config import EVENT_DAYS, TIME_BLOCKS
from models import Participant, Task, Assignment
import openpyxl
from rapidfuzz import fuzz

DUTCH_DAY_TO_INDEX = {
    "woensdag": 1,
    "donderdag": 2,
    "vrijdag": 3,
    "zaterdag": 4,
}

SURVEY_SLOT_TO_BLOCK = {
    "Mijn beschikbaarheid als hulpouder is: [07:00 - 07:30]": "07:00-07:30",
    "Mijn beschikbaarheid als hulpouder is: [07:30 - 09:00]": "07:30-09:00",
    "Mijn beschikbaarheid als hulpouder is: [09:00 - 13:00]": "09:00-13:00",
    "Mijn beschikbaarheid als hulpouder is: [13:00 - 15:30]": "13:00-15:30",
    "Mijn beschikbaarheid als hulpouder is: [15:30 - 18:00]": "15:30-18:00",
    "Mijn beschikbaarheid als hulpouder is: [18:00 - 21:00]": "18:00-21:00",
}

# Mapping of leader categories to task names
LEADER_CATEGORY_TO_TASKS = {
    "DINER": ["Avond eten uitdelen en opruimen"],
    "KOKEN": ["Avond eten uitdelen en opruimen"],
    "LUNCH": ["Lunch helpen", "Eten uitdelen en tafels opruimen"],
    "ONTBIJT": ["Ontbijt voorbereiden/opruimen"],
    "SCHOON": ["Schoonmaak"],
    "SNACK": ["fruitsnack", "namiddag snack", "avondsnack"],
}

# Normalize leader category aliases (maps variations to canonical category)
LEADER_CATEGORY_ALIASES = {
    "SCHOONMAAK": "SCHOON",
    "schoonmaak": "SCHOON",
    "Schoonmaak": "SCHOON",
    "schoon": "SCHOON",
    "FRUITSNACK": "SNACK",
    "fruitsnack": "SNACK",
    "Fruitsnack": "SNACK",
    "AVONDSNACK": "SNACK",
    "avondsnack": "SNACK",
    "Avondsnack": "SNACK",
    "NAMIDDAG SNACK": "SNACK",
    "namiddag snack": "SNACK",
    "Namiddag snack": "SNACK",
    "SNACK": "SNACK",
    "snack": "SNACK",
    "Snack": "SNACK",
}


def normalize_leader_category(category: str) -> str:
    """Normalize leader category to canonical form.
    
    Handles variations like:
    - "schoonmaak" / "SCHOON" -> "SCHOON"
    - "fruitsnack" / "avondsnack" / "namiddag snack" -> "SNACK"
    """
    if category in LEADER_CATEGORY_ALIASES:
        return LEADER_CATEGORY_ALIASES[category]
    return category.upper()


def parse_bool(value: str | None) -> bool:
    if value is None:
        return False
    if not isinstance(value, str):
        value = str(value)
    return value.strip().upper() in {"TRUE", "1", "YES", "Y", "JA", "J"}


def _safe_str(value) -> str | None:
    """Safely convert value to stripped string or None."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    # Convert non-string to string and strip
    stripped = str(value).strip()
    return stripped if stripped else None


def _normalize_phone(raw: str | None) -> str | None:
    """Normalize phone numbers: handle Excel dots, replace +358, handle missing leading zero."""
    if not raw:
        return None
    
    # Remove everything from the first dot onwards (common in Excel CSV exports)
    raw = str(raw).split(".")[0]
    
    # Remove dashes and spaces
    digits = raw.strip().replace(" ", "").replace("-", "")
    if not digits:
        return None
    # Replace +358 with 0 (Finnish format)
    if digits.startswith("+358"):
        digits = "0" + digits[4:]
    
    # If a phone number starts with a 4, assume a leading zero is missing
    if digits.startswith("4"):
        digits = "0" + digits

    return digits


def parse_days(value: str | None) -> set[int]:
    days: set[int] = set()
    if value is None:
        return days
    if not isinstance(value, str):
        value = str(value)
    for token in value.split(","):
        token_clean = token.strip().lower()
        if not token_clean:
            continue
        for dutch_name, day_index in DUTCH_DAY_TO_INDEX.items():
            if dutch_name in token_clean:
                days.add(day_index)
                break
    return days


def parse_group(value: str | None) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        value = str(value)
    group_value = value.strip().lower()
    if not group_value:
        return None
    # Remove both Dutch "groep" and English "group" prefixes
    group_value = group_value.replace("groep", "").replace("group", "").strip()
    if group_value in {"1", "2", "3a", "3b", "4", "5", "6+7", "8"}:
        return group_value
    # Normalize groep 6 or 7 to 6+7 (they're combined)
    if group_value in {"6", "7"}:
        return "6+7"
    return None


def parse_source_task_number(text: str | None) -> int | None:
    """Parse text like 'team van taak 4' or 'zelfde team als taak 4' to extract the source task number."""
    if not text:
        return None
    # Support "team van taak 4", "team taak 4", "zelfde team als taak 4", "taak 4 team"
    patterns = [
        r"team\s+(?:van\s+|als\s+)?taak\s+(\d+)",
        r"(?:zelfde\s+)?team\s+(?:als\s+)?taak\s+(\d+)",
        r"taak\s+(\d+)\s+team",
    ]
    for pattern in patterns:
        match = re.search(pattern, text.lower())
        if match:
            return int(match.group(1))
    return None


def merge_groups(existing_groups: str | None, new_group: str | None) -> str | None:
    """Merge group values, combining groups from multiple children into comma-separated list.
    
    For a parent with multiple kids, if existing_groups is "1" and new_group is "3b",
    returns "1,3b". Avoids duplicates and maintains order.
    """
    if not new_group:
        return existing_groups
    if not existing_groups:
        return new_group
    
    # Split existing groups and add new group if not already present
    existing_set = set(g.strip() for g in existing_groups.split(","))
    existing_set.add(new_group)
    
    # Return as comma-separated, maintaining a consistent order
    return ",".join(sorted(existing_set))


def merge_strings(existing: str | None, new_val: str | None, separator: str = ", ") -> str | None:
    """Merge two strings using a separator, avoiding duplicates."""
    if not new_val or not str(new_val).strip():
        return existing
    if not existing or not str(existing).strip():
        return new_val.strip()
    
    new_val = str(new_val).strip()
    existing = str(existing).strip()
    
    if new_val in existing:
        return existing
        
    return f"{existing}{separator}{new_val}"


def merge_availability(existing_map: dict[str, bool], new_map: dict[str, bool]) -> dict[str, bool]:
    """Merge availability from multiple children by combining them (OR operation).
    
    For a parent with multiple children, if available for any time slot for any child,
    mark as available for that slot.
    """
    result = {}
    for key in existing_map:
        result[key] = existing_map.get(key, False) or new_map.get(key, False)
    return result


def is_valid_email(email: str) -> bool:
    """Basic check to see if a string looks like an email address."""
    return "@" in email and "." in email


def parse_messaging(value: str | None) -> str:
    if value is None:
        return "whatsapp"
    if not isinstance(value, str):
        value = str(value)
    raw = value.strip().lower()
    if "signal" in raw:
        return "signal"
    if "telegram" in raw:
        return "telegram"
    if "sms" in raw:
        return "sms"
    if "whatsapp" in raw or "what'sapp" in raw or "whats app" in raw:
        return "whatsapp"
    return "whatsapp"


def find_participant_by_child_lastname(session, child_last: str | None) -> Participant | None:
    """Find a participant by matching child's last name.
    
    When multiple family members submit surveys separately, this helps identify
    that they belong to the same household. Returns the first matching participant
    with a child of the same last name.
    """
    if not child_last:
        return None
    child_last = child_last.strip()
    if not child_last:
        return None
    
    # Find participants with a child who has this last name
    participants = session.query(Participant).filter(
        Participant.child_last.ilike(f"%{child_last}%")
    ).all()
    
    # Return the first match (assumes they're ordered by insertion)
    if participants:
        return participants[0]
    return None


def get_parent_lastname(parent_name: str) -> str | None:
    """Extract the last name from a parent's full name.
    
    For "Jan Geertsma", returns "Geertsma".
    """
    if not parent_name:
        return None
    parts = parent_name.strip().split()
    if len(parts) > 1:
        return parts[-1]
    return None


def complete_child_name(child_first: str | None, child_last: str | None, parent_name: str) -> tuple[str | None, str | None]:
    """Complete child's name by adding parent's last name if child_last is missing.
    
    For child_first="Lorens", child_last=None, parent_name="Jan Geertsma", returns ("Lorens", "Geertsma").
    """
    if child_first and not child_last:
        parent_last = get_parent_lastname(parent_name)
        return (child_first, parent_last)
    return (child_first, child_last)


def complete_parent_name(parent_name: str, child_last: str | None) -> str:
    """Complete parent's name by adding child's last name if parent name is missing a last name.
    
    For parent_name="Jan", child_last="Geertsma", returns "Jan Geertsma".
    For parent_name="Jan Geertsma", child_last="Geertsma", returns "Jan Geertsma" (unchanged).
    """
    if not parent_name:
        return parent_name
    
    parts = parent_name.strip().split()
    # If parent name has only one word and child last name exists, append child's last name
    if len(parts) == 1 and child_last:
        return f"{parent_name.strip()} {child_last.strip()}"
    return parent_name


def _generate_name_variants(name: str) -> list[str]:
    """Generate various name formats for fuzzy matching.
    
    For "Jan Geertsma", generates:
    - "Jan Geertsma" (full name)
    - "Jan G" (first name + last initial)
    - "J Geertsma" (first initial + last name)
    - "Jan" (first name only)
    - "Geertsma" (last name only)
    """
    parts = name.strip().split()
    if len(parts) == 0:
        return [name]
    if len(parts) == 1:
        return [name]
    
    first_name = parts[0]
    last_name = parts[-1]
    first_initial = first_name[0].upper()
    last_initial = last_name[0].upper()
    
    variants = [
        name,  # Full name
        f"{first_name} {last_initial}",  # FirstName LastInitial
        f"{first_initial} {last_name}",  # FirstInitial LastName
        first_name,  # First name only
        last_name,  # Last name only
    ]
    
    # Remove duplicates while preserving order
    seen = set()
    unique_variants = []
    for v in variants:
        if v.lower() not in seen:
            seen.add(v.lower())
            unique_variants.append(v)
    
    return unique_variants


def build_survey_availability(row: dict[str, str]) -> dict[str, bool]:
    availability_fields = [f"day{day}_{block}" for day in range(1, EVENT_DAYS + 1) for block in TIME_BLOCKS]
    availability_map = {field: False for field in availability_fields}

    for slot_column, block in SURVEY_SLOT_TO_BLOCK.items():
        slot_value = row.get(slot_column)
        if slot_value is None:
            continue
        if not isinstance(slot_value, str):
            slot_value = str(slot_value)
        for day in parse_days(slot_value):
            availability_map[f"day{day}_{block}"] = True

    return availability_map


def is_survey_row(row: dict[str, str]) -> bool:
    return "Naam Ouder" in row and "E-mail Ouder" in row and "Voornaam Kind" in row


def parse_time(value: str) -> dt_time:
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(value.strip(), fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse time: {value!r}")


def infer_time_block(begin_time: dt_time, end_time: dt_time) -> str:
    """Infer the time block from the task start/end times."""
    midpoint_minutes = (begin_time.hour * 60 + begin_time.minute + end_time.hour * 60 + end_time.minute) / 2
    
    # 07:00 - 07:30 (420 - 450, mid 435)
    # 07:30 - 09:00 (450 - 540, mid 495)
    # 09:00 - 13:00 (540 - 780, mid 660)
    # 13:00 - 15:30 (780 - 930, mid 855)
    # 15:30 - 18:00 (930 - 1080, mid 1005)
    # 18:00 - 21:00 (1080 - 1260, mid 1170)
    
    if midpoint_minutes < 450: return "07:00-07:30"
    if midpoint_minutes < 540: return "07:30-09:00"
    if midpoint_minutes < 780: return "09:00-13:00"
    if midpoint_minutes < 930: return "13:00-15:30"
    if midpoint_minutes < 1080: return "15:30-18:00"
    return "18:00-21:00"


def _build_task_reader(handle: TextIO) -> csv.DictReader:
    return csv.DictReader(handle)


def _build_participant_reader(handle: TextIO) -> csv.DictReader:
    return csv.DictReader(handle)


def import_tasks_from_handle(handle: TextIO, session) -> dict:
    imported = 0
    skipped = 0
    updated = 0
    added = 0
    warnings: list[str] = []

    reader = _build_task_reader(handle)
    for row_number, row in enumerate(reader, start=2):
        try:
            task_name = row["task_name"].strip()
            day = int(row["day"].strip())
            begin_time = parse_time(row["begin_time"])
            end_time = parse_time(row["end_time"])
            points = int(float(row["points"].strip()))
            people_required = int(row["people_required"].strip())

            # Optional extra fields
            task_number = (
                int(row["task_number"].strip())
                if "task_number" in row and row["task_number"].strip()
                else None
            )
            description = row.get("description", "").strip() or None
            task_notes = row.get("task_notes", "").strip() or None
            lead_name = row.get("lead_name", "").strip() or None
            size = row.get("size", "").strip() or None
            location = row.get("location", "").strip() or None

            # Double points for 'Groot' tasks
            if size and size.strip().lower() == "groot":
                points *= 2
            
        except (KeyError, ValueError) as exc:
            warnings.append(f"tasks row {row_number} skipped - {exc}")
            skipped += 1
            continue

        time_block = infer_time_block(begin_time, end_time)
        source_task_number = parse_source_task_number(description) or parse_source_task_number(task_notes)

        # Look up by task_number first, then fallback
        existing = None
        if task_number is not None:
             existing = session.query(Task).filter_by(task_number=task_number).first()
        if existing is None:
             existing = session.query(Task).filter_by(name=task_name, day=day, begin_time=begin_time).first()

        if existing:
            existing.name = task_name
            existing.day = day
            existing.begin_time = begin_time
            existing.end_time = end_time
            existing.points = points
            existing.people_required = people_required
            existing.time_block = time_block
            existing.task_number = task_number
            existing.description = description
            existing.task_notes = task_notes
            existing.lead_name = lead_name
            existing.size = size
            existing.location = location
            existing.source_task_number = source_task_number
            updated += 1
        else:
            session.add(
                Task(
                    name=task_name,
                    day=day,
                    begin_time=begin_time,
                    end_time=end_time,
                    points=points,
                    people_required=people_required,
                    time_block=time_block,
                    task_number=task_number,
                    description=description,
                    task_notes=task_notes,
                    lead_name=lead_name,
                    size=size,
                    location=location,
                    source_task_number=source_task_number,
                )
            )
            added += 1
        imported += 1

    session.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "added": added,
        "updated": updated,
        "warnings": warnings,
    }


def import_participants_from_handle(handle: TextIO, session) -> dict:
    availability_fields = [f"day{day}_{block}" for day in range(1, EVENT_DAYS + 1) for block in TIME_BLOCKS]
    rows_processed = 0
    skipped = 0
    updated = 0
    added = 0
    warnings: list[str] = []

    reader = _build_participant_reader(handle)
    for row_number, row in enumerate(reader, start=2):
        try:
            if is_survey_row(row):
                name = row["Naam Ouder"].strip()
                email = row["E-mail Ouder"].strip().lower()
                phone = _normalize_phone(row.get("Telefoonnummer Ouder", ""))
                availability_map = build_survey_availability(row)
                child_notes = row.get("Opmerkingen:", "").strip() or None
                avail_notes = row.get("Opmerkingen beschikbaarheid:", "").strip() or None
                remarks_parts = [part for part in [child_notes, avail_notes] if part]
                remarks = " | ".join(remarks_parts) if remarks_parts else None
                
                # If parent name is missing but availability is provided, use child's last name
                if not name and any(availability_map.values()):
                    child_last = row.get("Achternaam Kind", "").strip()
                    if child_last:
                        name = child_last
            else:
                name = row["name"].strip()
                email = row["email"].strip().lower()
                phone = _normalize_phone(row.get("phone", ""))
                remarks = row.get("remarks", "").strip() or None
                availability_map = {field: parse_bool(row.get(field, "FALSE")) for field in availability_fields}

            if not name:
                child_info = ""
                if is_survey_row(row):
                    child_first = row.get("Voornaam Kind", "").strip() or "?"
                    child_last = row.get("Achternaam Kind", "").strip() or "?"
                    group = row.get("Groep", "").strip() or "?"
                    child_info = f" (Child: {child_first} {child_last}, Group: {group})"
                raise ValueError(f"missing parent name{child_info}")
            if not email:
                raise ValueError("missing email")
        except (KeyError, ValueError, TypeError, AttributeError) as exc:
            warnings.append(f"Row {row_number}: {str(exc)}")
            skipped += 1
            continue
        except Exception as exc:
            warnings.append(f"Row {row_number}: Unexpected error - {type(exc).__name__}: {str(exc)}")
            skipped += 1
            continue

        # 1. Try finding by email
        participant = session.query(Participant).filter_by(email=email).first()
        
        # 2. Try finding by name if not found by email (deduplication for user errors)
        if participant is None:
            participant = session.query(Participant).filter(Participant.name.ilike(name)).first()
            if participant:
                # If we found by name, check if we should update the email 
                # (e.g., if existing email is a phone number and new one is valid)
                if is_valid_email(email) and not is_valid_email(participant.email):
                    participant.email = email
        
        # 3. Try finding by child last name (for family members with same kids)
        if participant is None and is_survey_row(row):
            child_last = row.get("Achternaam Kind", "").strip() or None
            participant = find_participant_by_child_lastname(session, child_last)
            if participant:
                if is_valid_email(email) and not is_valid_email(participant.email):
                    # Update email if we found by child last name and new email is valid
                    participant.email = email

        is_new_participant = participant is None
        if participant is None:
            participant = Participant(name=name, email=email)
            session.add(participant)
            session.flush()
            added += 1
        else:
            updated += 1

        participant.name = name
        participant.phone = phone
        participant.remarks = merge_strings(participant.remarks, remarks, " | ")
        
        # Merge availability for existing participants (multiple children)
        if not is_new_participant:
            existing_availability = {
                f"day{day}_{block}": participant.get_block_availability(day, block)
                for day in range(1, EVENT_DAYS + 1)
                for block in TIME_BLOCKS
            }
            availability_map = merge_availability(existing_availability, availability_map)
        
        for field, available in availability_map.items():
            day_str, block = field.split("_")
            participant.set_block_availability(int(day_str.replace("day", "")), block, available)

        if is_survey_row(row):
            participant.submitted_at = row.get("Tijdstempel", "").strip() or None
            cf = row.get("Voornaam Kind", "").strip() or None
            cl = row.get("Achternaam Kind", "").strip() or None
            cf, cl = complete_child_name(cf, cl, name)
            
            participant.child_first = merge_strings(participant.child_first, cf)
            participant.child_last = merge_strings(participant.child_last, cl)
            
            # Complete parent name if missing last name
            name = complete_parent_name(name, cl)
            participant.name = name
            
            participant.group = merge_groups(participant.group, parse_group(row.get("Groep", "")))
            
            participant.child_att_d1 = merge_strings(participant.child_att_d1, row.get("Aanwezigheid van het kind [Woensdag]", "").strip() or None, " | ")
            participant.child_att_d2 = merge_strings(participant.child_att_d2, row.get("Aanwezigheid van het kind [Donderdag]", "").strip() or None, " | ")
            participant.child_att_d3 = merge_strings(participant.child_att_d3, row.get("Aanwezigheid van het kind [Vrijdag]", "").strip() or None, " | ")
            participant.child_att_d4 = merge_strings(participant.child_att_d4, row.get("Aanwezigheid van het kind [Zaterdag]", "").strip() or None, " | ")
            
            p_child_diet = row.get(
                "Geef hier eventuele allergieën of dieetwensen (bijv. vegetarisch, veganistisch) van je kind aan:",
                "",
            ).strip() or None
            participant.child_diet = merge_strings(participant.child_diet, p_child_diet, " | ")
            
            p_child_notes = row.get("Opmerkingen:", "").strip() or None
            participant.child_notes = merge_strings(participant.child_notes, p_child_notes, " | ")
            
            participant.first_ntc = participant.first_ntc or parse_bool(row.get("Dit is mijn eerste NTC zomerkamp", ""))
            
            p_sleep_notes = row.get("Opmerkingen overnachten:", "").strip() or None
            participant.sleep_notes = merge_strings(participant.sleep_notes, p_sleep_notes, " | ")
            
            p_avail_notes = row.get("Opmerkingen beschikbaarheid:", "").strip() or None
            participant.avail_notes = merge_strings(participant.avail_notes, p_avail_notes, " | ")
            
            participant.has_car = participant.has_car or parse_bool(
                row.get("Heb je een auto beschikbaar (om bijvoorbeeld leerkrachten van het station op te halen)", "")
            )
            
            p_parent_diet = row.get("Eventuele voedselallergieën Ouder:", "").strip() or None
            participant.parent_diet = merge_strings(participant.parent_diet, p_parent_diet, " | ")
            
            participant.survey_chat = row.get("Chat apps", "").strip() or None
            participant.messaging = parse_messaging(row.get("Chat apps", ""))

        rows_processed += 1

    try:
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise ValueError(f"Database integrity error: {exc.orig}") from exc

    # Calculate unique participants count (accounts for multiple children per adult)
    unique_participants = added + updated

    return {
        "imported": unique_participants,
        "rows_processed": rows_processed,
        "skipped": skipped,
        "added": added,
        "updated": updated,
        "warnings": warnings,
    }


def import_tasks_from_csv_path(csv_path: str, session) -> dict:
    with Path(csv_path).open(newline="", encoding="utf-8-sig") as handle:
        return import_tasks_from_handle(handle, session)


def import_participants_from_csv_path(csv_path: str, session) -> dict:
    with Path(csv_path).open(newline="", encoding="utf-8-sig") as handle:
        return import_participants_from_handle(handle, session)


def import_participants_from_excel_path(xlsx_path: str, session) -> dict:
    with open(xlsx_path, "rb") as f:
        return import_participants_from_excel(f.read(), session)


def import_participants_from_excel(file_bytes, session) -> dict:
    """Import participants from an Excel (.xlsx) workbook (bytes or file-like object).
    
    Supports both survey export format and simple CSV format.
    Deduplicates participants by email - if multiple children have the same parent email,
    they are treated as a single participant.
    """
    availability_fields = [f"day{day}_{block}" for day in range(1, EVENT_DAYS + 1) for block in TIME_BLOCKS]
    rows_processed = 0
    skipped = 0
    updated = 0
    added = 0
    warnings: list[str] = []

    if hasattr(file_bytes, "read"):
        file_bytes = file_bytes.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Extract headers from first row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {i: str(val).strip() if val else "" for i, val in enumerate(header_row)}
    
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_dict = {headers.get(i, f"col_{i}"): val for i, val in enumerate(row)}
            
            # Skip empty rows
            if not any(row_dict.values()):
                continue
            
            if is_survey_row(row_dict):
                name = _safe_str(row_dict.get("Naam Ouder")) or ""
                email = (_safe_str(row_dict.get("E-mail Ouder")) or "").lower()
                phone = _normalize_phone(_safe_str(row_dict.get("Telefoonnummer Ouder")))
                availability_map = build_survey_availability(row_dict)
                child_notes = _safe_str(row_dict.get("Opmerkingen:"))
                avail_notes = _safe_str(row_dict.get("Opmerkingen beschikbaarheid:"))
                remarks_parts = [part for part in [child_notes, avail_notes] if part]
                remarks = " | ".join(remarks_parts) if remarks_parts else None
                
                # If parent name is just a first name (no space) and we have child's last name,
                # combine them to create full parent name (e.g., "Jan" + "Geertsma" -> "Jan Geertsma")
                child_last = _safe_str(row_dict.get("Achternaam Kind"))
                if name and " " not in name and child_last:
                    # Parent name is likely just first name, append child's last name
                    name = f"{name} {child_last}"
                
                # If parent name is still missing but availability is provided, use child's last name
                if not name and any(availability_map.values()):
                    if child_last:
                        name = child_last
            else:
                name = _safe_str(row_dict.get("name")) or ""
                email = (_safe_str(row_dict.get("email")) or "").lower()
                phone = _normalize_phone(_safe_str(row_dict.get("phone")))
                remarks = _safe_str(row_dict.get("remarks"))
                availability_map = {field: parse_bool(_safe_str(row_dict.get(field))) for field in availability_fields}

            if not name:
                child_info = ""
                if is_survey_row(row_dict):
                    child_first = _safe_str(row_dict.get("Voornaam Kind")) or "?"
                    child_last = _safe_str(row_dict.get("Achternaam Kind")) or "?"
                    group = _safe_str(row_dict.get("Groep")) or "?"
                    child_info = f" (Child: {child_first} {child_last}, Group: {group})"
                raise ValueError(f"missing parent name{child_info}")
            if not email:
                raise ValueError("missing email")
        except (KeyError, ValueError, TypeError, AttributeError) as exc:
            warnings.append(f"Row {row_number}: {str(exc)}")
            skipped += 1
            continue
        except Exception as exc:
            warnings.append(f"Row {row_number}: Unexpected error - {type(exc).__name__}: {str(exc)}")
            skipped += 1
            continue

        # 1. Try finding by email
        participant = session.query(Participant).filter_by(email=email).first()
        
        # 2. Try finding by name if not found by email (deduplication for user errors)
        if participant is None:
            participant = session.query(Participant).filter(Participant.name.ilike(name)).first()
            if participant:
                # If we found by name, check if we should update the email 
                # (e.g., if existing email is a phone number and new one is valid)
                if is_valid_email(email) and not is_valid_email(participant.email):
                    participant.email = email
        
        # 3. Try finding by child last name (for family members with same kids)
        if participant is None and is_survey_row(row_dict):
            child_last = _safe_str(row_dict.get("Achternaam Kind"))
            participant = find_participant_by_child_lastname(session, child_last)
            if participant:
                if is_valid_email(email) and not is_valid_email(participant.email):
                    # Update email if we found by child last name and new email is valid
                    participant.email = email

        is_new_participant = participant is None
        if participant is None:
            participant = Participant(name=name, email=email)
            session.add(participant)
            session.flush()
            added += 1
        else:
            updated += 1

        participant.name = name
        participant.phone = phone
        participant.remarks = merge_strings(participant.remarks, remarks, " | ")
        
        # Merge availability for existing participants (multiple children)
        if not is_new_participant:
            existing_availability = {
                f"day{day}_{block}": participant.get_block_availability(day, block)
                for day in range(1, EVENT_DAYS + 1)
                for block in TIME_BLOCKS
            }
            availability_map = merge_availability(existing_availability, availability_map)
        
        for field, available in availability_map.items():
            day_str, block = field.split("_")
            participant.set_block_availability(int(day_str.replace("day", "")), block, available)

        participant.submitted_at = _safe_str(row_dict.get("Tijdstempel"))
        cf = _safe_str(row_dict.get("Voornaam Kind"))
        cl = _safe_str(row_dict.get("Achternaam Kind"))
        cf, cl = complete_child_name(cf, cl, name)
        
        participant.child_first = merge_strings(participant.child_first, cf)
        participant.child_last = merge_strings(participant.child_last, cl)
        
        # Complete parent name if missing last name
        name = complete_parent_name(name, cl)
        participant.name = name
        
        participant.group = merge_groups(participant.group, parse_group(_safe_str(row_dict.get("Groep"))))
        
        participant.child_att_d1 = merge_strings(participant.child_att_d1, _safe_str(row_dict.get("Aanwezigheid van het kind [Woensdag]")), " | ")
        participant.child_att_d2 = merge_strings(participant.child_att_d2, _safe_str(row_dict.get("Aanwezigheid van het kind [Donderdag]")), " | ")
        participant.child_att_d3 = merge_strings(participant.child_att_d3, _safe_str(row_dict.get("Aanwezigheid van het kind [Vrijdag]")), " | ")
        participant.child_att_d4 = merge_strings(participant.child_att_d4, _safe_str(row_dict.get("Aanwezigheid van het kind [Zaterdag]")), " | ")
        
        p_child_diet = _safe_str(row_dict.get("Geef hier eventuele allergieën of dieetwensen (bijv. vegetarisch, veganistisch) van je kind aan:"))
        participant.child_diet = merge_strings(participant.child_diet, p_child_diet, " | ")
        
        p_child_notes = _safe_str(row_dict.get("Opmerkingen:"))
        participant.child_notes = merge_strings(participant.child_notes, p_child_notes, " | ")
        
        participant.first_ntc = participant.first_ntc or parse_bool(_safe_str(row_dict.get("Dit is mijn eerste NTC zomerkamp")))
        
        p_sleep_notes = _safe_str(row_dict.get("Opmerkingen overnachten:"))
        participant.sleep_notes = merge_strings(participant.sleep_notes, p_sleep_notes, " | ")
        
        p_avail_notes = _safe_str(row_dict.get("Opmerkingen beschikbaarheid:"))
        participant.avail_notes = merge_strings(participant.avail_notes, p_avail_notes, " | ")
        
        participant.has_car = participant.has_car or parse_bool(_safe_str(row_dict.get("Heb je een auto beschikbaar (om bijvoorbeeld leerkrachten van het station op te halen)")))

        
        p_parent_diet = _safe_str(row_dict.get("Eventuele voedselallergieën Ouder:"))
        participant.parent_diet = merge_strings(participant.parent_diet, p_parent_diet, " | ")
        
        participant.survey_chat = _safe_str(row_dict.get("Chat apps"))
        participant.messaging = parse_messaging(_safe_str(row_dict.get("Chat apps"))) if _safe_str(row_dict.get("Chat apps")) else "whatsapp"

        rows_processed += 1

    try:
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise ValueError(f"Database integrity error: {exc.orig}") from exc

    # Calculate unique participants count (accounts for multiple children per adult)
    unique_participants = added + updated

    return {
        "imported": unique_participants,
        "rows_processed": rows_processed,
        "skipped": skipped,
        "added": added,
        "updated": updated,
        "warnings": warnings,
    }


def import_tasks_from_excel(file_bytes, session) -> dict:
    """Import tasks from an Excel (.xlsx) workbook (bytes or file-like object).

    Columns expected (Dutch headers):
        Datum, Start, Eind, Taak, Hoofdverantwoordelijke, Medewerkers,
        Informatie voor uitvoerenden, Grootte van taak, Locatie, Taak nummer,
        Aantekeningen voor Jan
    
    Column E (Hoofdverantwoordelijke) can contain leader names separated by:
        - Commas: "Name1, Name2, Name3"
        - Slashes: "Name1 / Name2 / Name3"
        - Mixed: "Name1, Name2 / Name3"
    Leaders will be automatically assigned to the task. "bestuur" is ignored.
    
    Column F (Medewerkers) can contain helper names and/or count of additional helpers:
        - Helper names: "Name1, Name2" or "Name1 / Name2"
        - Count of additional helpers: "2" or "3"
        - Mixed: "Name1, Name2, 2" (assigns Name1 & Name2 as helpers, then selects 2 more)
    """
    import openpyxl
    from models import Assignment

    imported = 0
    skipped = 0
    updated = 0
    added = 0
    warnings: list[str] = []

    if hasattr(file_bytes, "read"):
        file_bytes = file_bytes.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 11:
            warnings.append(f"row {row_number}: too few columns, skipped")
            skipped += 1
            continue

        datum, start, eind, taak, hoofdverantwoordelijke, aantal, info, grootte, locatie, taaknummer, aantekeningen = row[:11]

        # Skip section-header rows (e.g. "WOENSDAG") where Datum is None
        if datum is None:
            continue
        if not taak:
            skipped += 1
            continue

        try:
            # Map date to day index: Wednesday(weekday=2)→1, Thursday→2, Friday→3, Saturday→4
            if not hasattr(datum, "weekday"):
                warnings.append(f"row {row_number}: unexpected Datum type {type(datum).__name__}, skipped")
                skipped += 1
                continue
            day = datum.weekday() - 1  # Wednesday=2 → day 1
            if day < 1 or day > EVENT_DAYS:
                warnings.append(f"row {row_number}: date {datum.date()} maps to day {day} which is out of range 1-{EVENT_DAYS}, skipped")
                skipped += 1
                continue

            if not isinstance(start, dt_time):
                warnings.append(f"row {row_number}: invalid Start value {start!r}, skipped")
                skipped += 1
                continue
            if not isinstance(eind, dt_time):
                warnings.append(f"row {row_number}: invalid Eind value {eind!r}, skipped")
                skipped += 1
                continue

            task_name = str(taak).strip()

            # Determine leader count from Column E (Hoofdverantwoordelijke).
            # Column E may be: None, a number (0 or 1), a name, or multiple names
            # separated by commas/slashes.  "bestuur" entries are ignored.
            leader_count = 0
            if hoofdverantwoordelijke is not None:
                hoofd_str = str(hoofdverantwoordelijke).strip()
                if hoofd_str:
                    try:
                        leader_count = int(hoofd_str)
                    except ValueError:
                        hoofd_items = [
                            item.strip()
                            for item in hoofd_str.replace(" / ", ",").replace("/", ",").split(",")
                            if item.strip()
                        ]
                        leader_count = sum(
                            1 for item in hoofd_items
                            if not item.isdigit() and "bestuur" not in item.lower()
                        )

            # Determine worker count from Column F (aantal medewerkers).
            # Column F may be: None, a pure integer, names only, or names + count.
            helper_names = None
            worker_count = 0
            if aantal is not None:
                aantal_str = str(aantal).strip()
                try:
                    worker_count = int(aantal_str)
                except ValueError:
                    helper_names = aantal_str
                    raw_items = [
                        item.strip()
                        for item in aantal_str.replace(" / ", ",").replace("/", ",").split(",")
                        if item.strip()
                    ]
                    named = [
                        item for item in raw_items
                        if not item.isdigit() and "bestuur" not in item.lower()
                    ]
                    counts = [int(item) for item in raw_items if item.isdigit()]
                    worker_count = len(named) + (counts[-1] if counts else 0)

            # Total people required = leaders + workers (Column E + Column F)
            people_required = leader_count + worker_count
            
            task_number = int(taaknummer) if taaknummer is not None else None
            lead_name = str(hoofdverantwoordelijke).strip() if hoofdverantwoordelijke is not None else None
            description = str(info).strip() if info is not None else None
            size = str(grootte).strip() if grootte is not None else None
            location_val = str(locatie).strip() if locatie is not None else None
            task_notes_val = str(aantekeningen).strip() if aantekeningen is not None else None

        except (ValueError, TypeError) as exc:
            warnings.append(f"row {row_number} skipped - {exc}")
            skipped += 1
            continue

        time_block = infer_time_block(start, eind)
        source_task_number = parse_source_task_number(description) or parse_source_task_number(task_notes_val)

        # Double points for 'Groot' tasks
        points = 2 if size and size.strip().lower() == "groot" else 1

        # Look up by task_number first, then fall back to name/day/begin_time
        existing = None
        if task_number is not None:
            existing = session.query(Task).filter_by(task_number=task_number).first()
        if existing is None:
            existing = session.query(Task).filter_by(name=task_name, day=day, begin_time=start).first()

        if existing:
            existing.name = task_name
            existing.day = day
            existing.begin_time = start
            existing.end_time = eind
            existing.points = points
            existing.people_required = people_required
            existing.time_block = time_block
            existing.task_number = task_number
            existing.lead_name = lead_name
            existing.description = description
            existing.size = size
            existing.location = location_val
            existing.task_notes = task_notes_val
            existing.source_task_number = source_task_number
            updated += 1
            task = existing
        else:
            task = Task(
                name=task_name,
                day=day,
                begin_time=start,
                end_time=eind,
                points=points,
                people_required=people_required,
                time_block=time_block,
                task_number=task_number,
                lead_name=lead_name,
                description=description,
                size=size,
                location=location_val,
                task_notes=task_notes_val,
                source_task_number=source_task_number,
            )
            session.add(task)
            added += 1
        
        session.flush()
        
        # Parse leader names from column E and create assignments
        if lead_name:
            # Split by both comma and slash, handle mixed separators
            # Replace slashes with commas first, then split by comma
            leader_names_raw = lead_name.replace(" / ", ",").replace("/", ",")
            leader_names = [name.strip() for name in leader_names_raw.split(",") if name.strip()]
            
            # Get all participants for matching
            all_participants = session.query(Participant).all()
            
            # First, delete any existing non-reserved assignments for this task
            # (we're replacing with reserved assignments from Column E)
            session.query(Assignment).filter(
                Assignment.task_id == task.id,
                Assignment.is_reserved == False
            ).delete()
            session.flush()
            
            for leader_name_item in leader_names:
                # Skip numeric values (they represent number of leaders required, not names)
                if leader_name_item.isdigit():
                    continue
                
                # Skip "bestuur" and similar administrative entries
                if "bestuur" in leader_name_item.lower():
                    continue
                
                # Find matching participant using fuzzy matching
                matching_participant = None
                best_score = 0
                
                leader_variants = _generate_name_variants(leader_name_item)
                for participant in all_participants:
                    participant_variants = _generate_name_variants(participant.name)
                    
                    for leader_variant in leader_variants:
                        for participant_variant in participant_variants:
                            score = fuzz.ratio(leader_variant.lower(), participant_variant.lower())
                            if score > best_score:
                                best_score = score
                                matching_participant = participant
                
                if matching_participant and best_score >= 60:
                    # Check if already assigned (reserved)
                    existing_assignment = session.query(Assignment).filter_by(
                        task_id=task.id,
                        participant_id=matching_participant.id,
                        is_reserved=True
                    ).first()
                    
                    if not existing_assignment:
                        session.add(Assignment(
                            task_id=task.id,
                            participant_id=matching_participant.id,
                            role="lead",
                            points_awarded=task.points,
                            is_reserved=True,
                        ))
                else:
                    if best_score < 60:
                        warnings.append(f"row {row_number}: leader '{leader_name_item}' not found (fuzzy match score: {best_score})")
        
        # Parse helper names from column F and create assignments
        # Helper assignment: handle both named helpers and numeric count
        all_participants = session.query(Participant).all()
        named_helpers = []
        additional_helper_count = 0
        has_explicit_names = False
        
        if helper_names:
            # Split by both comma and slash, handle mixed separators
            helper_names_raw = helper_names.replace(" / ", ",").replace("/", ",")
            helper_items = [item.strip() for item in helper_names_raw.split(",") if item.strip()]
            
            # Extract numeric values (count of additional helpers) and filter them out
            for item in helper_items:
                if item.isdigit():
                    # Use the last numeric value as the additional helper count
                    additional_helper_count = int(item)
                else:
                    # Skip "bestuur" and similar administrative entries
                    if "bestuur" not in item.lower():
                        named_helpers.append(item)
            
            # Check if we have explicit named helpers
            has_explicit_names = len(named_helpers) > 0
            
            # Assign named helpers (mark as reserved since explicitly requested)
            for helper_name_item in named_helpers:
                # Find matching participant using fuzzy matching
                matching_participant = None
                best_score = 0
                
                helper_variants = _generate_name_variants(helper_name_item)
                for participant in all_participants:
                    participant_variants = _generate_name_variants(participant.name)
                    
                    for helper_variant in helper_variants:
                        for participant_variant in participant_variants:
                            score = fuzz.ratio(helper_variant.lower(), participant_variant.lower())
                            if score > best_score:
                                best_score = score
                                matching_participant = participant
                
                if matching_participant and best_score >= 60:
                    # Check if already assigned to this task
                    existing_assignment = session.query(Assignment).filter_by(
                        task_id=task.id,
                        participant_id=matching_participant.id
                    ).first()
                    
                    if not existing_assignment:
                        session.add(Assignment(
                            task_id=task.id,
                            participant_id=matching_participant.id,
                            role="helper",
                            points_awarded=task.points,
                            is_reserved=True,
                        ))
                else:
                    if best_score < 60:
                        warnings.append(f"row {row_number}: helper '{helper_name_item}' not found (fuzzy match score: {best_score})")
            
            # Select additional helpers randomly
            # If there are explicit names AND a count, fill that many more
            # If Column F is purely numeric (no names), don't assign random helpers here -
            # let the scheduler handle it to maintain flexibility
            if has_explicit_names and additional_helper_count > 0:
                # Get list of participants already assigned to this task
                already_assigned_ids = set(
                    assignment.participant_id 
                    for assignment in session.query(Assignment).filter_by(task_id=task.id).all()
                )
                
                # Get available candidates: those not yet assigned and available for the time slot
                available_candidates = [
                    p for p in all_participants 
                    if p.id not in already_assigned_ids
                    and p.get_block_availability(day, time_block)
                ]
                
                # Randomly select up to additional_helper_count from available candidates
                # NOTE: These are NOT marked as reserved, so scheduler can rebalance them
                helpers_to_add = min(additional_helper_count, len(available_candidates))
                if helpers_to_add > 0:
                    selected_helpers = random.sample(available_candidates, helpers_to_add)
                    
                    for helper_participant in selected_helpers:
                        session.add(Assignment(
                            task_id=task.id,
                            participant_id=helper_participant.id,
                            role="helper",
                            points_awarded=task.points,
                            is_reserved=False,  # NOT reserved - scheduler can reassign
                        ))
        
        imported += 1

    session.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "added": added,
        "updated": updated,
        "warnings": warnings,
    }


def import_tasks_from_excel_path(xlsx_path: str, session) -> dict:
    with open(xlsx_path, "rb") as f:
        return import_tasks_from_excel(f.read(), session)


def find_matching_task(target_task_name: str, all_tasks: list) -> tuple[object | None, int]:
    """Find a task matching the target name using multiple strategies.
    
    Tries in order:
    1. Exact case-insensitive match
    2. Substring match (target is substring of task name)
    3. Fuzzy matching with lower threshold
    
    Returns (matching_task, score) or (None, 0) if not found.
    """
    
    target_lower = target_task_name.lower()
    
    # Strategy 1: Exact case-insensitive match
    for task in all_tasks:
        if task.name.lower() == target_lower:
            return (task, 100)
    
    # Strategy 2: Substring match (target is contained in task name)
    for task in all_tasks:
        task_lower = task.name.lower()
        if target_lower in task_lower or task_lower in target_lower:
            return (task, 90)
    
    # Strategy 3: Fuzzy matching
    matching_task = None
    best_score = 0
    for task in all_tasks:
        score = fuzz.ratio(target_lower, task.name.lower())
        if score > best_score:
            best_score = score
            matching_task = task
    
    if matching_task and best_score >= 60:  # Lowered threshold from 70 to 60
        return (matching_task, best_score)
    
    return (None, best_score)


def import_leaders_from_handle(handle: TextIO, session) -> dict:
    """Import leaders from a text file with format: TASK_NAME: name1, name2, name3
    
    Uses fuzzy matching to find participants and tasks by name (handles partial/missing names).
    Creates Assignment records with role='lead' for matching participants and tasks.
    """
    
    imported = 0
    skipped = 0
    updated = 0
    added = 0
    warnings: list[str] = []
    
    # Read all lines and parse task: name1, name2, ... format
    for line_number, line in enumerate(handle, start=1):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        
        # Parse "TASK_NAME: name1, name2, name3"
        if ":" not in line:
            warnings.append(f"line {line_number}: missing ':' separator")
            skipped += 1
            continue
        
        task_part, names_part = line.split(":", 1)
        task_name = task_part.strip()
        # Normalize leader category (handles variations like "schoonmaak" -> "SCHOON")
        task_name = normalize_leader_category(task_name)
        leader_names = [name.strip() for name in names_part.split(",") if name.strip()]
        
        if not task_name:
            warnings.append(f"line {line_number}: missing task name")
            skipped += 1
            continue
        
        if not leader_names:
            warnings.append(f"line {line_number}: no leader names found")
            skipped += 1
            continue
        
        # Get list of task names to match (from mapping or fuzzy match)
        matching_tasks = []
        
        # Check if this is a known category with mapped task names
        if task_name in LEADER_CATEGORY_TO_TASKS:
            task_names_to_find = LEADER_CATEGORY_TO_TASKS[task_name]
        else:
            task_names_to_find = [task_name]
        
        # Find matching tasks
        all_tasks = session.query(Task).all()
        
        for target_task_name in task_names_to_find:
            matching_task, match_score = find_matching_task(target_task_name, all_tasks)
            
            if matching_task:
                matching_tasks.append(matching_task)
            else:
                warnings.append(f"line {line_number}: task '{target_task_name}' not found (best fuzzy match score: {match_score})")
        
        if not matching_tasks:
            warnings.append(f"line {line_number}: no tasks found for category '{task_name}'")
            skipped += 1
            continue
        
        # Process each leader name
        all_participants = session.query(Participant).all()
        
        for leader_name in leader_names:
            # Find matching participant using fuzzy matching
            # Try multiple name formats: full name, FirstName LastInitial, FirstInitial LastName, etc.
            matching_participant = None
            best_participant_score = 0
            
            leader_name_variants = _generate_name_variants(leader_name)
            
            for participant in all_participants:
                participant_name_variants = _generate_name_variants(participant.name)
                
                # Score each combination of leader name variant vs participant name variant
                for leader_variant in leader_name_variants:
                    for participant_variant in participant_name_variants:
                        score = fuzz.ratio(leader_variant.lower(), participant_variant.lower())
                        if score > best_participant_score:
                            best_participant_score = score
                            matching_participant = participant
            
            if matching_participant is None or best_participant_score < 60:
                warnings.append(f"line {line_number}: leader '{leader_name}' not found (fuzzy match score: {best_participant_score})")
                skipped += 1
                continue
            
            # Set is_leader flag to mark them as experienced for task leadership preference
            if not matching_participant.is_leader:
                matching_participant.is_leader = True
                updated += 1
            
            imported += 1

    try:
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise ValueError(f"Database integrity error: {exc.orig}") from exc
    
    return {
        "imported": imported,
        "skipped": skipped,
        "added": added,
        "updated": updated,
        "warnings": warnings,
    }


def import_leaders_from_txt_path(txt_path: str, session) -> dict:
    with Path(txt_path).open(newline="", encoding="utf-8-sig") as handle:
        return import_leaders_from_handle(handle, session)
